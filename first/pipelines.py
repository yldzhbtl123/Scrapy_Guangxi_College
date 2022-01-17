# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import openpyxl
import os


class AnquangongchengPipeline:
    def process_item(self, item, spider):
        print(item)
        if not os.path.exists('spider.xlsx'):
            wb = openpyxl.Workbook()
            wb.save('spider.xlsx')
        # 获取工作薄的活动表
        wb = openpyxl.load_workbook('spider.xlsx')
        add_all_anquangongcheng_trans = item['add_all_anquangongcheng_trans']
        sheet = wb.active
        for i in add_all_anquangongcheng_trans:
            sheet.append(i)  # 每次写入一行
        wb.save('spider.xlsx')  # 记得保存格式为xlsx

        return item

class ShuilidianliPipeline:
    def process_item(self, item, spider):
        if not os.path.exists('spider.xlsx'):
            wb = openpyxl.Workbook()
            wb.save('spider.xlsx')
        # 获取工作薄的活动表
        wb = openpyxl.load_workbook('spider.xlsx')
        sheet = wb.active
        add_all_shuili_trans= item['add_all_shuili_trans']
        for i in add_all_shuili_trans:
            sheet.append(i)# 每次写入一行
        wb.save('spider.xlsx')  # 记得保存格式为xlsx

        return item

class GongshangPipeline:
    def process_item(self, item, spider):
        if not os.path.exists('spider.xlsx'):
            wb = openpyxl.Workbook()
            wb.save('spider.xlsx')
        # 获取工作薄的活动表
        wb = openpyxl.load_workbook('spider.xlsx')
        sheet = wb.active
        q =[]
        print(item)
        if "招聘" in item["title"]:
            for key,value in item.items():
                print(key,value)
                q.append(value)
            sheet.append(q)
            print(q)
            q.clear()
        else:
            pass
        wb.save('spider.xlsx')  # 记得保存格式为xlsx

        return item

class ZiranziyuanPipeline:
    def process_item(self, item, spider):
        if not os.path.exists('spider.xlsx'):
            wb = openpyxl.Workbook()
            wb.save('spider.xlsx')
        # 获取工作薄的活动表
        wb = openpyxl.load_workbook('spider.xlsx')
        sheet = wb.active
        add_all_ziranziyuan_trans = item['add_all_ziranziyuan_trans']
        for i in add_all_ziranziyuan_trans:
            print(i)
            sheet.append(i)# 每次写入一行
        wb.save('spider.xlsx')  # 记得保存格式为xlsx

        return item

class ShengtaizhiyePipeline:
    def process_item(self, item, spider):
        if not os.path.exists('spider.xlsx'):
            wb = openpyxl.Workbook()
            wb.save('spider.xlsx')
        # 获取工作薄的活动表
        wb = openpyxl.load_workbook('spider.xlsx')
        sheet = wb.active
        q =[]
        if "招聘" in item["title"]:
            for key,value in item.items():
                q.append(value)
            sheet.append(q)
            q.clear()
        else:
            pass
        wb.save('spider.xlsx')  # 记得保存格式为xlsx

        return item

class JidianPipeline:
    def process_item(self, item, spider):
        if not os.path.exists('spider.xlsx'):
            wb = openpyxl.Workbook()
            wb.save('spider.xlsx')
        # 获取工作薄的活动表
        wb = openpyxl.load_workbook('spider.xlsx')
        sheet = wb.active
        q =[]
        if "招聘" in item["title"]:
            for key,value in item.items():
                q.append(value)
            sheet.append(q)
            q.clear()
        else:
            pass
        wb.save('spider.xlsx')  # 记得保存格式为xlsx

        return item